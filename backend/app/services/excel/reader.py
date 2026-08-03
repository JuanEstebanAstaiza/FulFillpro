from __future__ import annotations

import io
import unicodedata
from datetime import date, datetime
from typing import Any, Optional

from openpyxl import load_workbook

from backend.app.config import get_settings


def norm(s: Any) -> str:
    s = str(s or "").strip().upper()
    s = unicodedata.normalize("NFD", s)
    return "".join(c for c in s if not unicodedata.combining(c))


COL_ALIASES = {
    "id": ["ID", "ORDEN", "ORDER ID", "ID ORDEN"],
    "guia": ["NUMERO GUIA", "GUIA", "N GUIA", "TRACKING", "NUMERO DE GUIA", "N° GUIA"],
    "producto": ["PRODUCTO", "PRODUCT", "NOMBRE PRODUCTO", "ITEM", "ARTICULO"],
    "variacion": ["VARIACION", "VARIATION", "TALLA", "COLOR"],
    "cantidad": ["CANTIDAD", "QTY", "QUANTITY", "CANT", "UNIDADES"],
    "valor": ["TOTAL DE LA ORDEN", "TOTAL ORDEN", "VALOR", "PRECIO", "TOTAL"],
    "fechaGuia": [
        "FECHA GUIA GENERADA",
        "FECHA GUIA",
        "FECHA DE ENVIO",
        "SHIP DATE",
        "FECHA GUIA GEN",
    ],
    "ciudad": [
        "CIUDAD",
        "CIUDAD DESTINO",
        "MUNICIPIO",
        "CITY",
        "DESTINO",
        "CIUDAD DE DESTINO",
    ],
    "transportadora": [
        "TRANSPORTADORA",
        "TRANSPORTADOR",
        "CARRIER",
        "COURIER",
        "EMPRESA TRANSPORTE",
        "LOGISTICA",
        "MENSAJERIA",
    ],
}


def read_excel_rows(content: bytes) -> list[dict[str, Any]]:
    settings = get_settings()
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    it = ws.iter_rows(values_only=True)
    raw_headers = next(it, None)
    if not raw_headers:
        wb.close()
        raise ValueError("El archivo no tiene encabezados.")
    headers = [str(h or "").strip() for h in raw_headers]

    idx: dict[str, int] = {}
    for key, aliases in COL_ALIASES.items():
        wanted = [norm(a) for a in aliases]
        for i, h in enumerate(headers):
            if norm(h) in wanted:
                idx[key] = i
                break

    faltan = [n for k, n in (("producto", "PRODUCTO"), ("guia", "NUMERO GUIA")) if k not in idx]
    if faltan:
        wb.close()
        raise ValueError(
            f"No se encontraron las columnas: {', '.join(faltan)}. "
            f"Columnas del archivo: {', '.join(headers[:15])}"
        )

    rows: list[dict[str, Any]] = []
    for raw in it:
        if not any(v for v in raw if v is not None):
            continue
        if len(rows) >= settings.max_rows:
            wb.close()
            raise ValueError(
                f"El archivo supera el límite de {settings.max_rows:,} órdenes. Divide el reporte en partes."
            )

        def get(key: str, default: str = "") -> str:
            i = idx.get(key)
            if i is None or i >= len(raw):
                return default
            val = raw[i]
            # Preservar IDs largos como texto
            if isinstance(val, float) and val == int(val):
                return str(int(val))
            return str(val if val is not None else "").strip()

        try:
            cant = int(float(get("cantidad", "1") or "1"))
        except Exception:
            cant = 1
        try:
            val = float(get("valor", "0") or "0")
        except Exception:
            val = 0.0

        producto = get("producto")
        guia = get("guia")
        if not producto and not guia:
            continue

        rows.append(
            {
                "id": get("id"),
                "guia": guia,
                "producto": producto,
                "variacion": get("variacion"),
                "cantidad": cant,
                "valor": val,
                "fechaGuia": get("fechaGuia"),
                "ciudad": get("ciudad"),
                "transportadora": get("transportadora"),
            }
        )
    wb.close()
    return rows


def parse_date(s: str) -> Optional[date]:
    if not s:
        return None
    text = str(s).strip()
    # openpyxl a veces deja datetime
    if isinstance(s, datetime):
        return s.date()
    if isinstance(s, date) and not isinstance(s, datetime):
        return s
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%Y/%m/%d", "%d/%m/%y"):
        try:
            return datetime.strptime(text[:10], fmt).date()
        except Exception:
            pass
    return None
