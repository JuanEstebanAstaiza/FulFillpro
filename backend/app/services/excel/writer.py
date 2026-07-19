from __future__ import annotations

import io
from datetime import date, datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

P = {
    "VC": "1B5E20",
    "VM": "2E7D32",
    "VL": "E8F5E9",
    "GH": "263238",
    "GL": "ECEFF1",
    "RH": "B71C1C",
    "RL": "FFEBEE",
    "NL": "FFF3E0",
    "NH": "E65100",
    "AL": "FFF8E1",
    "BL": "E3F2FD",
}

_THIN = Side(style="thin", color="B0BEC5")
_BRD = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_SM = Side(style="medium", color=P["GH"])

_fills: dict = {}
_fonts: dict = {}
_aligns: dict = {}


def _fill(h: str) -> PatternFill:
    if h not in _fills:
        _fills[h] = PatternFill("solid", start_color=h)
    return _fills[h]


def _font(h: str, sz: int, b: bool) -> Font:
    k = (h, sz, b)
    if k not in _fonts:
        _fonts[k] = Font(name="Calibri", color=h, size=sz, bold=b)
    return _fonts[k]


def _align(ah: str, wrap: bool) -> Alignment:
    k = (ah, wrap)
    if k not in _aligns:
        _aligns[k] = Alignment(horizontal=ah, vertical="center", wrap_text=wrap)
    return _aligns[k]


def sty(c, bg, fg="1A1A1A", sz=10, b=False, ah="left", wrap=True):
    c.font = _font(fg, sz, b)
    c.fill = _fill(bg)
    c.alignment = _align(ah, wrap)
    c.border = _BRD


def hdr(c, bg, ah="center"):
    sty(c, bg, "FFFFFF", 10, True, ah, True)


def banner(ws, ncols, text, bg, size, height):
    ws.append([text])
    r = ws.max_row
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    c = ws.cell(r, 1)
    c.font = _font("FFFFFF", size, True)
    c.fill = _fill(bg)
    c.alignment = _align("left", False)
    ws.row_dimensions[r].height = height


def subtitle(ws, ncols, text):
    ws.append([text])
    r = ws.max_row
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    c = ws.cell(r, 1)
    c.font = _font("455A64", 10, False)
    c.fill = _fill("F5F5F5")
    c.alignment = _align("left", False)
    ws.row_dimensions[r].height = 18


def spacer(ws, ncols):
    ws.append([])
    r = ws.max_row
    for ci in range(1, ncols + 1):
        ws.cell(r, ci).fill = _fill("FAFAFA")
    ws.row_dimensions[r].height = 6


def company_strip(ws, ncols: int, company_name: str, company_code: str, license_code: str):
    """Franja distintiva de la empresa contratante (anti reventa / uso ajeno)."""
    label = f"LICENCIADO PARA: {company_name or 'CLIENTE'}"
    if company_code:
        label += f"  ·  CÓDIGO EMPRESA: {company_code}"
    if license_code:
        label += f"  ·  LIC. {license_code}"
    label += "  ·  USO EXCLUSIVO — NO TRANSFERIBLE"
    ws.append([label])
    r = ws.max_row
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    c = ws.cell(r, 1)
    c.font = _font("FFFFFF", 9, True)
    c.fill = _fill("0D3B12")
    c.alignment = _align("left", False)
    ws.row_dimensions[r].height = 20


def brand_footer(ws, ncols: int, text: str):
    ws.append([text])
    r = ws.max_row
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    c = ws.cell(r, 1)
    c.font = Font(name="Calibri", color="546E7A", size=8, italic=True)
    c.alignment = _align("left", True)
    ws.row_dimensions[r].height = 28


def apply_print_brand(ws, company_name: str, license_code: str):
    ws.oddHeader.center.text = f"FulfillPro · {company_name}"
    ws.oddHeader.center.font = "Calibri,Bold"
    ws.oddHeader.center.size = 12
    ws.oddFooter.left.text = f"Lic. {license_code} · Uso exclusivo {company_name}"
    ws.oddFooter.center.text = "Página &P de &N"
    ws.oddFooter.right.text = "Confidencial"


def build_excel(
    resumen_final: list[dict[str, Any]],
    cant_cols: list[str],
    cant_max: int,
    reporte: list[dict],
    prior: list[dict],
    total_riesgo: float,
    today: date,
    company_name: str = "",
    company_code: str = "",
    license_code: str = "",
) -> bytes:
    wb = Workbook()
    today_str = today.strftime("%d/%m/%Y")
    now_str = datetime.now().strftime("%d/%m/%Y %H:%M")
    co = company_name or "Cliente"
    footer_txt = (
        f"Generado por FulfillPro · {now_str} · Documento exclusivo para {co} "
        f"({company_code or '—'}) · Licencia {license_code or '—'} · "
        f"Prohibida la reventa o uso por otras empresas"
    )

    # Unidades físicas a alistar: sumatoria (tamaño_paquete × n_órdenes) por fila
    total_uds = sum(int(row.get("TOTAL_UNIDADES", 0) or 0) for row in resumen_final)
    n_combos = sum(1 for r in resumen_final if str(r.get("VARIABLES", "")).upper() == "COMBO")
    # VARIACION | PRODUCTO | Cant.1..N | TOTAL UNID.
    ncols1 = 2 + cant_max + 1
    total_col = ncols1  # última columna

    # HOJA 1: RESUMEN
    # Filas: 1 banner, 2 company, 3 metrics, 4 spacer, 5 headers, 6+ data
    ws1 = wb.active
    ws1.title = "Resumen"
    ws1.sheet_properties.tabColor = P["VC"]
    ws1.sheet_view.showGridLines = False
    ws1.column_dimensions["A"].width = 12
    ws1.column_dimensions["B"].width = 38
    for i in range(3, ncols1):
        ws1.column_dimensions[get_column_letter(i)].width = 9
    ws1.column_dimensions[get_column_letter(total_col)].width = 12

    banner(ws1, ncols1, f"FulfillPro · {co} — Resumen de Órdenes para Bodega", P["VC"], 14, 30)
    company_strip(ws1, ncols1, co, company_code, license_code)
    subtitle(
        ws1,
        ncols1,
        f"{len(resumen_final)} productos - {total_uds} unidades a alistar - {n_combos} combos | Fecha: {today_str}",
    )
    spacer(ws1, ncols1)

    ws1.append(
        ["VARIACION", "PRODUCTO"]
        + [f"Cant. {i}" for i in range(1, cant_max + 1)]
        + ["TOTAL UNID."]
    )
    r = ws1.max_row
    header_row = r
    data_start = header_row + 1
    ws1.row_dimensions[r].height = 26
    hdr(ws1.cell(r, 1), P["GH"])
    hdr(ws1.cell(r, 2), P["GH"], "left")
    for ci in range(3, ncols1):
        hdr(ws1.cell(r, ci), P["VC"])
    hdr(ws1.cell(r, total_col), "0D47A1")  # azul: total unidades a despachar
    ws1.freeze_panes = f"A{data_start}"

    for i, row in enumerate(resumen_final):
        es_c = str(row.get("VARIABLES", "")).upper() == "COMBO"
        bg = P["AL"] if es_c else (P["VL"] if i % 2 == 0 else P["GL"])
        vals = [row.get("VARIABLES", ""), row.get("PRODUCTO", "")] + [
            row.get(f"Cantidad {j}", "") or None for j in range(1, cant_max + 1)
        ] + [int(row.get("TOTAL_UNIDADES", 0) or 0) or None]
        ws1.append(vals)
        r = ws1.max_row
        ws1.row_dimensions[r].height = 30
        sty(
            ws1.cell(r, 1),
            bg,
            "8D6E63" if es_c else (P["VC"] if row.get("VARIABLES") else "78909C"),
            10,
            es_c,
            "center",
            True,
        )
        sty(ws1.cell(r, 2), bg, "4E342E" if es_c else "212121", 10, es_c, "left", True)
        for ci in range(3, ncols1):
            if ws1.cell(r, ci).value:
                sty(ws1.cell(r, ci), bg, "5D4037" if es_c else P["VC"], 11, True, "center", False)
            else:
                sty(ws1.cell(r, ci), bg, "CFD8DC", 10, False, "center", False)
        # TOTAL UNID. resaltado
        c_tot = ws1.cell(r, total_col)
        if c_tot.value:
            sty(c_tot, "E3F2FD", "0D47A1", 12, True, "center", False)
        else:
            sty(c_tot, bg, "CFD8DC", 10, False, "center", False)

    sep = ws1.max_row + 1
    ws1.row_dimensions[sep].height = 4
    for ci in range(1, ncols1 + 1):
        c = ws1.cell(sep, ci)
        c.fill = _fill(P["GH"])
        c.border = Border(top=_SM, bottom=_SM)

    data_end = data_start + len(resumen_final) - 1 if resumen_final else data_start
    total_letter = get_column_letter(total_col)
    ws1.append(
        [""]
        + [f"TOTAL - {total_uds} unidades"]
        + [
            f"=SUM({get_column_letter(ci + 3)}{data_start}:{get_column_letter(ci + 3)}{data_end})"
            for ci in range(cant_max)
        ]
        + [f"=SUM({total_letter}{data_start}:{total_letter}{data_end})"]
    )
    r = ws1.max_row
    ws1.row_dimensions[r].height = 26
    ws1.cell(r, 1).fill = _fill(P["VM"])
    ws1.cell(r, 1).border = _BRD
    hdr(ws1.cell(r, 2), P["VM"], "left")
    for ci in range(3, ncols1):
        hdr(ws1.cell(r, ci), P["VM"])
    hdr(ws1.cell(r, total_col), "0D47A1")

    brand_footer(ws1, ncols1, footer_txt)

    ws1.page_setup.orientation = "portrait"
    ws1.page_setup.paperSize = 9
    ws1.page_setup.fitToPage = True
    ws1.page_setup.fitToWidth = 1
    ws1.page_setup.fitToHeight = 0
    ws1.print_title_rows = f"1:{header_row}"
    ws1.page_margins = PageMargins(left=0.5, right=0.5, top=0.75, bottom=0.75)
    apply_print_brand(ws1, co, license_code)

    # HOJA 2: REPORTE ORDENADO
    ws2 = wb.create_sheet("Reporte Ordenado")
    ws2.sheet_properties.tabColor = P["GH"]
    ws2.sheet_view.showGridLines = False
    ws2.column_dimensions["A"].width = 14
    ws2.column_dimensions["B"].width = 38
    ws2.column_dimensions["C"].width = 10

    banner(ws2, 3, f"FulfillPro · {co} — Reporte Ordenado", P["VC"], 13, 26)
    company_strip(ws2, 3, co, company_code, license_code)
    subtitle(ws2, 3, f"{len(reporte)} lineas | Fecha: {today_str}")
    spacer(ws2, 3)

    ws2.sheet_format.defaultRowHeight = 20
    ws2.append(["ID ORDEN", "PRODUCTO", "CANTIDAD"])
    r = ws2.max_row
    ws2.row_dimensions[r].height = 24
    hdr(ws2.cell(r, 1), P["GH"])
    hdr(ws2.cell(r, 2), P["GH"], "left")
    hdr(ws2.cell(r, 3), P["GH"])
    ws2.freeze_panes = f"A{r + 1}"

    prev = ""
    for i, row in enumerate(reporte):
        is_n = row["PRODUCTO"] != prev
        bg = P["BL"] if i % 2 == 0 else P["GL"]
        ws2.append([str(row["ID ORDEN"]), row["PRODUCTO"], int(row["CANTIDAD"] or 0)])
        r = ws2.max_row
        sty(ws2.cell(r, 1), bg, "546E7A", 9, False, "center", False)
        sty(ws2.cell(r, 2), bg, "212121", 10, is_n, "left", True)
        sty(ws2.cell(r, 3), bg, "1565C8", 11, True, "center", False)
        prev = row["PRODUCTO"]

    brand_footer(ws2, 3, footer_txt)
    apply_print_brand(ws2, co, license_code)

    # HOJA 3: PRIORITARIAS
    ws3 = wb.create_sheet("PRIORITARIAS")
    ws3.sheet_properties.tabColor = P["RH"]
    ws3.sheet_view.showGridLines = False
    for i, w in enumerate([18, 34, 12, 13, 13, 18, 13], 1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    banner(ws3, 7, f"FulfillPro · {co} — Órdenes Prioritarias", P["RH"], 13, 28)
    company_strip(ws3, 7, co, company_code, license_code)
    subtitle(ws3, 7, f"{len(prior)} ordenes atrasadas - Riesgo: ${total_riesgo:,.0f} | Fecha: {today_str}")
    subtitle(ws3, 7, "Rojo intenso = 5+ dias | Rojo suave = 2-4 dias | Naranja = 1 dia")

    ws3.sheet_format.defaultRowHeight = 24
    ws3.append(["N GUIA", "PRODUCTO", "VALOR", "FECHA GUIA", "DIAS RETRASO", "ESTADO", "RIESGO 20%"])
    r = ws3.max_row
    prior_header = r
    prior_data_start = r + 1
    ws3.row_dimensions[r].height = 26
    for ci in range(1, 8):
        hdr(ws3.cell(r, ci), P["RH"])
    hdr(ws3.cell(r, 2), P["RH"], "left")
    ws3.freeze_panes = f"A{prior_data_start}"

    if not prior:
        ws3.append(["Sin órdenes atrasadas para la fecha de hoy.", "", "", "", "", "", ""])
        r = ws3.max_row
        ws3.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
        c = ws3.cell(r, 1)
        c.font = Font(name="Calibri", color="90A4AE", size=10, italic=True)

    for row in prior:
        dias = int(row["DIAS RETRASO"])
        if dias >= 5:
            bg, fg = "FFCDD2", P["RH"]
        elif dias >= 2:
            bg, fg = P["RL"], "C62828"
        else:
            bg, fg = P["NL"], P["NH"]
        ws3.append(
            [
                str(row["N GUIA"]),
                row["PRODUCTO"],
                float(row["VALOR"]),
                str(row["FECHA GUIA"]),
                dias,
                row["ESTADO"],
                float(row["RIESGO 20"]),
            ]
        )
        r = ws3.max_row
        sty(ws3.cell(r, 1), bg, "546E7A", 9, False, "center", False)
        sty(ws3.cell(r, 2), bg, fg, 10, True, "left", True)
        c3 = ws3.cell(r, 3)
        sty(c3, bg, "37474F", 10, False, "right", False)
        c3.number_format = "$#,##0"
        sty(ws3.cell(r, 4), bg, "546E7A", 9, False, "center", False)
        sty(ws3.cell(r, 5), bg, fg, 13, True, "center", False)
        sty(ws3.cell(r, 6), bg, fg, 9, dias >= 2, "center", False)
        c7 = ws3.cell(r, 7)
        sty(c7, bg, P["RH"], 10, True, "right", False)
        c7.number_format = "$#,##0"

    if prior:
        sep = ws3.max_row + 1
        for ci in range(1, 8):
            c = ws3.cell(sep, ci)
            c.fill = _fill(P["GH"])
            c.border = Border(top=_SM, bottom=_SM)
        ws3.row_dimensions[sep].height = 4
        lp = prior_data_start + len(prior) - 1
        ws3.append(["", "", "", "", "", "TOTAL RIESGO:", f"=SUM(G{prior_data_start}:G{lp})"])
        r = ws3.max_row
        ws3.row_dimensions[r].height = 26
        for ci in range(1, 6):
            c = ws3.cell(r, ci)
            c.fill = _fill(P["RH"])
            c.border = _BRD
        hdr(ws3.cell(r, 6), P["RH"], "right")
        c = ws3.cell(r, 7)
        hdr(c, P["RH"], "right")
        c.number_format = "$#,##0"

    brand_footer(ws3, 7, footer_txt)
    apply_print_brand(ws3, co, license_code)

    # Propiedades del libro con marca de empresa
    wb.properties.title = f"FulfillPro — {co}"
    wb.properties.subject = f"Licencia {license_code} · {company_code}"
    wb.properties.creator = f"FulfillPro ({co})"
    wb.properties.keywords = f"fulfillpro,{company_code},{license_code}"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
