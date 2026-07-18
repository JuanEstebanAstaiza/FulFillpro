from fastapi import FastAPI, UploadFile, File, Form, HTTPException, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import Optional
from sqlalchemy import create_engine, Column, String, Integer, Boolean, DateTime, Text, JSON
from sqlalchemy.orm import declarative_base, sessionmaker
from contextlib import contextmanager
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from datetime import datetime, date, timedelta
import os, io, re, unicodedata, threading, urllib.request, time

# ══════════════════════════════════════════════════════════
# CONFIGURACION
# ══════════════════════════════════════════════════════════
ADMIN_SECRET      = os.environ.get("ADMIN_SECRET", "CAMBIA-ESTO-EN-RENDER")
DEVICE_STALE_DAYS = int(os.environ.get("DEVICE_STALE_DAYS", "60"))
MAX_ROWS          = int(os.environ.get("MAX_ROWS", "60000"))
MAX_CANT_COLS     = int(os.environ.get("MAX_CANT_COLS", "60"))
TYPE_DEVICES      = {"standard": 1, "pro": 3, "enterprise": 999}

# ══════════════════════════════════════════════════════════
# BASE DE DATOS
# ══════════════════════════════════════════════════════════
DATABASE_URL = os.environ.get("DATABASE_URL", "sqlite:///./fulfillpro.db")
if DATABASE_URL.startswith("postgres://"):
    DATABASE_URL = DATABASE_URL.replace("postgres://", "postgresql://", 1)

engine = create_engine(
    DATABASE_URL,
    connect_args={"check_same_thread": False} if "sqlite" in DATABASE_URL else {},
    pool_pre_ping=True, pool_recycle=300,
)
SessionLocal = sessionmaker(bind=engine, autocommit=False, autoflush=False)
Base = declarative_base()

class License(Base):
    __tablename__ = "licenses"
    code        = Column(String, primary_key=True)
    label       = Column(String, default="")
    type        = Column(String, default="standard")
    max_devices = Column(Integer, default=1)
    uses        = Column(Integer, default=0)
    limit_uses  = Column(Integer, default=0)
    expiry      = Column(String, default="")
    devices     = Column(JSON, default=list)
    active      = Column(Boolean, default=True)
    created_at  = Column(DateTime, default=datetime.utcnow)
    last_access = Column(DateTime, nullable=True)

class AccessLog(Base):
    __tablename__ = "access_log"
    id          = Column(Integer, primary_key=True, autoincrement=True)
    code        = Column(String, index=True)
    label       = Column(String, default="")
    tipo        = Column(String)
    detalle     = Column(Text, default="")
    fingerprint = Column(String, default="")
    ip          = Column(String, default="")
    time        = Column(DateTime, default=datetime.utcnow)

Base.metadata.create_all(bind=engine)

@contextmanager
def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

def seed():
    with get_db() as db:
        if db.query(License).count() == 0:
            db.add(License(code="DEMO-001", label="Demo interno", type="standard",
                           max_devices=1, limit_uses=20, expiry="2027-12-31", devices=[]))
            db.commit()
seed()

# ══════════════════════════════════════════════════════════
# KEEP-ALIVE
# ══════════════════════════════════════════════════════════
def _keep_alive():
    url = os.environ.get("RENDER_SERVICE_URL", "")
    if not url:
        return
    while True:
        time.sleep(14 * 60)
        try:
            urllib.request.urlopen(f"{url}/health", timeout=10)
        except Exception:
            pass

threading.Thread(target=_keep_alive, daemon=True).start()

# ══════════════════════════════════════════════════════════
# APP
# ══════════════════════════════════════════════════════════
app = FastAPI(docs_url=None, redoc_url=None)
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_credentials=False,
                   allow_methods=["*"], allow_headers=["*"])

class ValidateRequest(BaseModel):
    code: str
    device_fingerprint: str          # ID estable guardado en el equipo
    device_name: str = ""
    device_soft: str = ""            # huella de respaldo (datos que no cambian)

class AdminAction(BaseModel):
    admin_secret: str
    action: str
    code: str = ""
    data: Optional[dict] = {}

def days_left(expiry):
    if not expiry:
        return 9999
    try:
        return (datetime.strptime(expiry, "%Y-%m-%d").date() - date.today()).days
    except Exception:
        return 9999

def log_event(db, code, label, tipo, detalle, fp="", ip=""):
    db.add(AccessLog(code=code, label=label, tipo=tipo, detalle=detalle,
                     fingerprint=fp, ip=ip))
    db.commit()

# ══════════════════════════════════════════════════════════
# DISPOSITIVOS  (acepta formato viejo y nuevo)
# ══════════════════════════════════════════════════════════
def norm_devices(raw):
    out = []
    for d in (raw or []):
        if isinstance(d, str):
            # Formato antiguo: huella suelta, sin metadatos -> se considera "legacy"
            out.append({"id": d, "soft": "", "name": "", "first_seen": "", "last_seen": "", "legacy": True})
        elif isinstance(d, dict) and d.get("id"):
            out.append({
                "id":         d.get("id", ""),
                "soft":       d.get("soft", ""),
                "name":       d.get("name", ""),
                "first_seen": d.get("first_seen", ""),
                "last_seen":  d.get("last_seen", ""),
                "legacy":     bool(d.get("legacy", False)),
            })
    return out

def _age_days(iso):
    if not iso:
        return 99999
    try:
        return (datetime.utcnow() - datetime.fromisoformat(iso)).days
    except Exception:
        return 99999

def resolve_device(devices, dev_id, soft, name):
    """Devuelve (devices_actualizados, estado). Estado: known | rebound | new | replaced | full"""
    now = datetime.utcnow().isoformat(timespec="seconds")

    # 1) El mismo ID ya registrado
    for d in devices:
        if dev_id and d["id"] == dev_id:
            d["last_seen"] = now
            d["legacy"] = False
            if soft: d["soft"] = soft
            if name: d["name"] = name
            return devices, "known"

    # 2) Mismo equipo con ID perdido (borraron cache) -> re-vincular, NO gastar cupo
    if soft:
        for d in devices:
            if d.get("soft") and d["soft"] == soft:
                d["id"] = dev_id
                d["last_seen"] = now
                d["legacy"] = False
                if name: d["name"] = name
                return devices, "rebound"

    nuevo = {"id": dev_id, "soft": soft, "name": name,
             "first_seen": now, "last_seen": now, "legacy": False}

    # 3) Hay cupo libre
    if len(devices) < MAXD[0]:
        devices.append(nuevo)
        return devices, "new"

    # 4) Sin cupo: reciclar entradas viejas (legacy) o inactivas
    reciclables = [d for d in devices if d.get("legacy") or _age_days(d.get("last_seen")) >= DEVICE_STALE_DAYS]
    if reciclables:
        viejo = sorted(reciclables, key=lambda d: _age_days(d.get("last_seen")))[-1]
        devices.remove(viejo)
        devices.append(nuevo)
        return devices, "replaced"

    return devices, "full"

MAXD = [1]  # contenedor mutable para pasar el limite a resolve_device

# ══════════════════════════════════════════════════════════
# ENDPOINT 1 — VALIDAR LICENCIA
# ══════════════════════════════════════════════════════════
@app.post("/api/validate")
async def validate_license(req: ValidateRequest, request: Request):
    code = req.code.upper().strip()
    dev  = (req.device_fingerprint or "").strip()
    soft = (req.device_soft or "").strip()
    name = (req.device_name or "").strip()
    ip   = request.client.host

    if not dev:
        raise HTTPException(400, "No se pudo identificar el dispositivo.")

    with get_db() as db:
        lic = db.query(License).filter_by(code=code).first()
        if not lic:
            raise HTTPException(400, "Codigo de licencia no reconocido.")
        if not lic.active:
            log_event(db, code, lic.label, "Bloqueado", "Licencia desactivada", dev, ip)
            raise HTTPException(403, "Licencia desactivada. Contacta al administrador.")
        if days_left(lic.expiry) <= 0:
            log_event(db, code, lic.label, "Rechazado", f"Expirada ({lic.expiry})", dev, ip)
            raise HTTPException(403, f"Licencia expirada el {lic.expiry}.")
        if lic.limit_uses > 0 and lic.uses >= lic.limit_uses:
            log_event(db, code, lic.label, "Rechazado", "Usos agotados", dev, ip)
            raise HTTPException(403, "Limite de procesamientos alcanzado.")

        MAXD[0] = lic.max_devices or 1
        devices, estado = resolve_device(norm_devices(lic.devices), dev, soft, name)

        if estado == "full":
            log_event(db, code, lic.label, "Bloqueado",
                      f"Cupo lleno ({lic.max_devices}). Intento desde {name} / {dev}", dev, ip)
            raise HTTPException(403,
                f"Esta licencia ya esta vinculada a {lic.max_devices} dispositivo(s). "
                f"Pide al administrador liberar un dispositivo o ampliar el plan.")

        lic.devices = devices          # reasignar: SQLAlchemy no detecta cambios in-place en JSON
        lic.last_access = datetime.utcnow()

        etiqueta = {
            "known":    "Acceso OK",
            "rebound":  "Re-vinculado",
            "replaced": "Dispositivo reemplazado",
            "new":      "Nuevo dispositivo",
        }[estado]
        log_event(db, code, lic.label, etiqueta, f"{name} ({dev})", dev, ip)
        db.commit()

        return {
            "ok": True,
            "label": lic.label,
            "type": lic.type,
            "device_status": estado,
            "devices_used": len(devices),
            "devices_max": lic.max_devices,
            "days_left": days_left(lic.expiry),
            "expiry": lic.expiry,
            "uses": lic.uses,
            "limit": lic.limit_uses,
        }

# ══════════════════════════════════════════════════════════
# LECTURA DEL EXCEL
# ══════════════════════════════════════════════════════════
def norm(s):
    s = str(s or "").strip().upper()
    s = unicodedata.normalize("NFD", s)
    return "".join(c for c in s if not unicodedata.combining(c))

COL_ALIASES = {
    "id":        ["ID", "ORDEN", "ORDER ID", "ID ORDEN"],
    "guia":      ["NUMERO GUIA", "GUIA", "N GUIA", "TRACKING", "NUMERO DE GUIA"],
    "producto":  ["PRODUCTO", "PRODUCT", "NOMBRE PRODUCTO", "ITEM", "ARTICULO"],
    "variacion": ["VARIACION", "VARIATION", "TALLA", "COLOR"],
    "cantidad":  ["CANTIDAD", "QTY", "QUANTITY", "CANT", "UNIDADES"],
    "valor":     ["TOTAL DE LA ORDEN", "TOTAL ORDEN", "VALOR", "PRECIO", "TOTAL"],
    "fechaGuia": ["FECHA GUIA GENERADA", "FECHA GUIA", "FECHA DE ENVIO", "SHIP DATE", "FECHA GUIA GEN"],
}

def read_excel_rows(content):
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    it = ws.iter_rows(values_only=True)
    raw_headers = next(it, None)
    if not raw_headers:
        wb.close()
        raise ValueError("El archivo no tiene encabezados.")
    headers = [str(h or "").strip() for h in raw_headers]

    idx = {}
    for key, aliases in COL_ALIASES.items():
        wanted = [norm(a) for a in aliases]
        for i, h in enumerate(headers):
            if norm(h) in wanted:
                idx[key] = i
                break

    faltan = [n for k, n in (("producto", "PRODUCTO"), ("guia", "NUMERO GUIA")) if k not in idx]
    if faltan:
        wb.close()
        raise ValueError(f"No se encontraron las columnas: {', '.join(faltan)}. "
                         f"Columnas del archivo: {', '.join(headers[:15])}")

    rows = []
    for raw in it:
        if not any(v for v in raw if v is not None):
            continue
        if len(rows) >= MAX_ROWS:
            wb.close()
            raise ValueError(f"El archivo supera el limite de {MAX_ROWS:,} ordenes. Divide el reporte en partes.")

        def get(key, default=""):
            i = idx.get(key)
            return str(raw[i] or "").strip() if i is not None and i < len(raw) else default

        try:    cant = int(float(get("cantidad", "1") or "1"))
        except Exception: cant = 1
        try:    val = float(get("valor", "0") or "0")
        except Exception: val = 0.0

        rows.append({"id": get("id"), "guia": get("guia"), "producto": get("producto"),
                     "variacion": get("variacion"), "cantidad": cant, "valor": val,
                     "fechaGuia": get("fechaGuia")})
    wb.close()
    return rows

def parse_date(s):
    if not s:
        return None
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%Y/%m/%d"):
        try:
            return datetime.strptime(s.strip(), fmt).date()
        except Exception:
            pass
    return None

# ══════════════════════════════════════════════════════════
# PROCESAMIENTO  (misma logica de siempre)
# ══════════════════════════════════════════════════════════
def process_rows(rows, today):
    rows.sort(key=lambda r: r["producto"])

    by_guia = {}
    for r in rows:
        if r["guia"]:
            by_guia.setdefault(r["guia"], []).append(r)

    combo_guias = {g for g, items in by_guia.items() if len(items) > 1}
    resumen_dict, nombre_dict = {}, {}
    cant_max = 1

    for guia, items in by_guia.items():
        if guia in combo_guias:
            id_final = f"COMP-{guia}"
            cant_res = 1
            nombre   = " + ".join(f"{r['producto']} ({r['cantidad']})" for r in items)
        else:
            r = items[0]
            var = r["variacion"]
            id_final = f"{r['id']}|{var}" if var and var not in ("nan", "") else r["id"]
            cant_res = max(r["cantidad"], 1)
            nombre   = r["producto"]

        # Guarda de seguridad: evita generar cientos de columnas por una cantidad atipica
        cant_res = min(cant_res, MAX_CANT_COLS)

        if id_final not in resumen_dict:
            resumen_dict[id_final] = {}
            nombre_dict[id_final]  = nombre
        resumen_dict[id_final][cant_res] = resumen_dict[id_final].get(cant_res, 0) + 1
        cant_max = max(cant_max, cant_res)

    unified = {}
    for key, cnts in resumen_dict.items():
        if "|" in key:                          variable = key.split("|", 1)[1]
        elif key.startswith("COMP-"):           variable = "COMBO"
        elif re.fullmatch(r"\d+(\.\d+)?", key): variable = ""
        else:                                   variable = key

        prod = nombre_dict[key]
        if variable != "COMBO":
            prod = re.sub(r"\s*\(\d+\)\s*", " ", prod).strip()

        ukey = f"{variable}|{prod}"
        if ukey not in unified:
            unified[ukey] = {"VARIABLES": variable, "PRODUCTO": prod,
                             **{f"Cantidad {c}": 0 for c in range(1, cant_max + 1)}}
        for c, n in cnts.items():
            unified[ukey][f"Cantidad {c}"] = unified[ukey].get(f"Cantidad {c}", 0) + n

    resumen_final = sorted(unified.values(), key=lambda r: r["PRODUCTO"])
    for row in resumen_final:
        for c in range(1, cant_max + 1):
            v = row.get(f"Cantidad {c}", 0)
            row[f"Cantidad {c}"] = int(v) if v and int(v) > 0 else ""

    reporte = sorted(
        [{"ID ORDEN": r["id"], "PRODUCTO": r["producto"], "CANTIDAD": r["cantidad"]} for r in rows],
        key=lambda r: (r["PRODUCTO"], r["CANTIDAD"]),
    )

    prior = []
    for r in rows:
        fg = parse_date(r["fechaGuia"])
        if fg:
            dias = (today - fg).days
            if dias >= 1:
                prior.append({"N GUIA": r["guia"], "PRODUCTO": r["producto"], "VALOR": r["valor"],
                              "FECHA GUIA": str(fg), "DIAS RETRASO": dias,
                              "ESTADO": "URGENTE" if dias == 1 else "SUPER ATRASADA",
                              "RIESGO 20": round(r["valor"] * 0.2)})
    prior.sort(key=lambda r: -r["DIAS RETRASO"])

    return resumen_final, cant_max, reporte, prior, sum(r["RIESGO 20"] for r in prior)

# ══════════════════════════════════════════════════════════
# ENDPOINT 2 — PROCESAR
# ══════════════════════════════════════════════════════════
@app.post("/api/process")
async def process_excel(request: Request,
                        file: UploadFile = File(...),
                        code: str = Form(""),
                        device_fingerprint: str = Form(""),
                        device_soft: str = Form("")):
    code = code.upper().strip()
    dev  = (device_fingerprint or "").strip()
    soft = (device_soft or "").strip()

    with get_db() as db:
        lic = db.query(License).filter_by(code=code).first()
        if not lic or not lic.active:
            raise HTTPException(403, "Licencia no valida.")
        if days_left(lic.expiry) <= 0:
            raise HTTPException(403, "Licencia expirada.")
        if lic.limit_uses > 0 and lic.uses >= lic.limit_uses:
            raise HTTPException(403, "Limite de procesamientos alcanzado.")

        devices = norm_devices(lic.devices)
        autorizado = any(d["id"] == dev for d in devices) or \
                     (soft and any(d.get("soft") == soft for d in devices))
        if not autorizado:
            raise HTTPException(403, "Dispositivo no autorizado.")

    content = await file.read()
    try:
        rows = read_excel_rows(content)
    except ValueError as e:
        raise HTTPException(400, str(e))
    except Exception as e:
        raise HTTPException(400, f"Error leyendo archivo: {e}")

    if not rows:
        raise HTTPException(400, "El archivo no tiene datos.")

    today = date.today()
    try:
        resumen_final, cant_max, reporte, prior, total_riesgo = process_rows(rows, today)
        cant_cols = [f"Cantidad {c}" for c in range(1, cant_max + 1)]
        output = build_excel(resumen_final, cant_cols, cant_max, reporte, prior, total_riesgo, today)
    except Exception as e:
        raise HTTPException(500, f"Error procesando datos: {e}")

    # El contador de usos cuenta PROCESAMIENTOS, no inicios de sesion
    with get_db() as db:
        lic = db.query(License).filter_by(code=code).first()
        if lic:
            lic.uses = (lic.uses or 0) + 1
            db.commit()
            log_event(db, code, lic.label, "Proceso",
                      f"{len(rows)} ordenes procesadas", dev, request.client.host)

    return StreamingResponse(
        io.BytesIO(output),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=FulfillPro_{today}.xlsx"},
    )

# ══════════════════════════════════════════════════════════
# ENDPOINT 3 — ADMIN
# ══════════════════════════════════════════════════════════
@app.post("/api/admin")
async def admin_action(req: AdminAction):
    if req.admin_secret != ADMIN_SECRET:
        raise HTTPException(403, "No autorizado.")

    with get_db() as db:
        if req.action == "list":
            lics = []
            for l in db.query(License).order_by(License.created_at).all():
                lics.append({
                    "code": l.code, "label": l.label, "type": l.type,
                    "max_devices": l.max_devices, "uses": l.uses, "limit": l.limit_uses,
                    "expiry": l.expiry, "devices": norm_devices(l.devices),
                    "active": l.active, "days_left": days_left(l.expiry),
                    "last_access": l.last_access.isoformat() if l.last_access else None,
                })
            logs = [{"code": x.code, "label": x.label, "tipo": x.tipo, "detalle": x.detalle,
                     "fingerprint": x.fingerprint, "ip": x.ip, "time": x.time.isoformat()}
                    for x in db.query(AccessLog).order_by(AccessLog.time.desc()).limit(200).all()]
            return {"licenses": lics, "log": logs}

        if req.action == "release_devices":
            lic = db.query(License).filter_by(code=req.code).first()
            if not lic: raise HTTPException(404, "No existe.")
            lic.devices = []
            log_event(db, req.code, lic.label, "Admin", "Dispositivos liberados")
            db.commit()
            return {"ok": True}

        if req.action == "release_one":
            lic = db.query(License).filter_by(code=req.code).first()
            if not lic: raise HTTPException(404, "No existe.")
            target = req.data.get("device_id", "")
            lic.devices = [d for d in norm_devices(lic.devices) if d["id"] != target]
            log_event(db, req.code, lic.label, "Admin", f"Dispositivo liberado: {target}")
            db.commit()
            return {"ok": True}

        if req.action == "set_devices":
            lic = db.query(License).filter_by(code=req.code).first()
            if not lic: raise HTTPException(404, "No existe.")
            lic.max_devices = max(1, int(req.data.get("max_devices", 1)))
            log_event(db, req.code, lic.label, "Admin", f"Cupo de dispositivos: {lic.max_devices}")
            db.commit()
            return {"ok": True, "max_devices": lic.max_devices}

        if req.action == "renew":
            lic = db.query(License).filter_by(code=req.code).first()
            if not lic: raise HTTPException(404, "No existe.")
            d = int(req.data.get("days", 30))
            base = datetime.strptime(lic.expiry, "%Y-%m-%d").date() \
                   if lic.expiry and days_left(lic.expiry) > 0 else date.today()
            lic.expiry = (base + timedelta(days=d)).strftime("%Y-%m-%d")
            log_event(db, req.code, lic.label, "Admin", f"Renovada hasta {lic.expiry}")
            db.commit()
            return {"ok": True, "new_expiry": lic.expiry}

        if req.action == "add":
            d = req.data
            c = d.get("code", "").upper().strip()
            if not c: raise HTTPException(400, "Codigo requerido.")
            if db.query(License).filter_by(code=c).first():
                raise HTTPException(400, "Ese codigo ya existe.")
            t = d.get("type", "standard")
            db.add(License(code=c, label=d.get("label", ""), type=t,
                           max_devices=TYPE_DEVICES.get(t, 1),
                           limit_uses=int(d.get("limit", 0)),
                           expiry=d.get("expiry", ""), devices=[], active=True))
            db.commit()
            return {"ok": True}

        if req.action == "delete":
            lic = db.query(License).filter_by(code=req.code).first()
            if lic:
                db.delete(lic); db.commit()
            return {"ok": True}

        if req.action == "toggle":
            lic = db.query(License).filter_by(code=req.code).first()
            if not lic: raise HTTPException(404, "No existe.")
            lic.active = not lic.active
            db.commit()
            return {"ok": True, "active": lic.active}

        if req.action == "reset_uses":
            lic = db.query(License).filter_by(code=req.code).first()
            if lic:
                lic.uses = 0; db.commit()
            return {"ok": True}

    raise HTTPException(400, "Accion no reconocida.")

@app.get("/health")
async def health():
    with get_db() as db:
        n = db.query(License).count()
    return {"ok": True, "licenses": n}

@app.get("/")
async def root():
    return {"status": "FulfillPro API activa"}

# ══════════════════════════════════════════════════════════
# GENERADOR DE EXCEL  (mismo diseno, con cache de estilos)
# ══════════════════════════════════════════════════════════
P = {"VC": "1B5E20", "VM": "2E7D32", "VL": "E8F5E9", "GH": "263238", "GL": "ECEFF1",
     "RH": "B71C1C", "RL": "FFEBEE", "NL": "FFF3E0", "NH": "E65100",
     "AL": "FFF8E1", "BL": "E3F2FD"}

_THIN = Side(style="thin", color="B0BEC5")
_BRD  = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_SM   = Side(style="medium", color=P["GH"])

_fills, _fonts, _aligns = {}, {}, {}

def _fill(h):
    if h not in _fills:
        _fills[h] = PatternFill("solid", start_color=h)
    return _fills[h]

def _font(h, sz, b):
    k = (h, sz, b)
    if k not in _fonts:
        _fonts[k] = Font(name="Calibri", color=h, size=sz, bold=b)
    return _fonts[k]

def _align(ah, wrap):
    k = (ah, wrap)
    if k not in _aligns:
        _aligns[k] = Alignment(horizontal=ah, vertical="center", wrap_text=wrap)
    return _aligns[k]

def sty(c, bg, fg="1A1A1A", sz=10, b=False, ah="left", wrap=True):
    c.font = _font(fg, sz, b); c.fill = _fill(bg)
    c.alignment = _align(ah, wrap); c.border = _BRD

def hdr(c, bg, ah="center"):
    sty(c, bg, "FFFFFF", 10, True, ah, True)

def banner(ws, ncols, text, bg, size, height):
    ws.append([text])
    r = ws.max_row
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    c = ws.cell(r, 1)
    c.font = _font("FFFFFF", size, True); c.fill = _fill(bg)
    c.alignment = _align("left", False)
    ws.row_dimensions[r].height = height

def subtitle(ws, ncols, text):
    ws.append([text])
    r = ws.max_row
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    c = ws.cell(r, 1)
    c.font = _font("455A64", 10, False); c.fill = _fill("F5F5F5")
    c.alignment = _align("left", False)
    ws.row_dimensions[r].height = 18

def spacer(ws, ncols):
    ws.append([])
    r = ws.max_row
    for ci in range(1, ncols + 1):
        ws.cell(r, ci).fill = _fill("FAFAFA")
    ws.row_dimensions[r].height = 6

def build_excel(resumen_final, cant_cols, cantMax, reporte, prior, total_riesgo, today):
    wb = Workbook()
    todayStr = today.strftime("%d/%m/%Y")
    nowStr   = datetime.now().strftime("%d/%m/%Y %H:%M")

    total_uds = sum(int(row.get(c, 0) or 0) for row in resumen_final for c in cant_cols if row.get(c, "") != "")
    n_combos  = sum(1 for r in resumen_final if str(r.get("VARIABLES", "")).upper() == "COMBO")
    ncols1    = 2 + cantMax

    # ── HOJA 1: RESUMEN ──
    ws1 = wb.active
    ws1.title = "Resumen"
    ws1.sheet_properties.tabColor = P["VC"]
    ws1.sheet_view.showGridLines = False
    ws1.column_dimensions["A"].width = 12
    ws1.column_dimensions["B"].width = 38
    for i in range(3, ncols1 + 1):
        ws1.column_dimensions[get_column_letter(i)].width = 9

    banner(ws1, ncols1, "FulfillPro - Resumen de Ordenes para Bodega", P["VC"], 14, 30)
    subtitle(ws1, ncols1, f"{len(resumen_final)} productos - {total_uds} unidades - {n_combos} combos | Fecha: {todayStr}")
    spacer(ws1, ncols1)

    ws1.append(["VARIACION", "PRODUCTO"] + [f"Cant. {i}" for i in range(1, cantMax + 1)])
    r = ws1.max_row
    ws1.row_dimensions[r].height = 26
    hdr(ws1.cell(r, 1), P["GH"])
    hdr(ws1.cell(r, 2), P["GH"], "left")
    for ci in range(3, ncols1 + 1):
        hdr(ws1.cell(r, ci), P["VC"])
    ws1.freeze_panes = "A5"

    for i, row in enumerate(resumen_final):
        esC = str(row.get("VARIABLES", "")).upper() == "COMBO"
        bg  = P["AL"] if esC else (P["VL"] if i % 2 == 0 else P["GL"])
        vals = [row.get("VARIABLES", ""), row.get("PRODUCTO", "")] + \
               [row.get(f"Cantidad {j}", "") or None for j in range(1, cantMax + 1)]
        ws1.append(vals)
        r = ws1.max_row
        ws1.row_dimensions[r].height = 30
        sty(ws1.cell(r, 1), bg, "8D6E63" if esC else (P["VC"] if row.get("VARIABLES") else "78909C"), 10, esC, "center", True)
        sty(ws1.cell(r, 2), bg, "4E342E" if esC else "212121", 10, esC, "left", True)
        for ci in range(3, ncols1 + 1):
            if ws1.cell(r, ci).value:
                sty(ws1.cell(r, ci), bg, "5D4037" if esC else P["VC"], 11, True, "center", False)
            else:
                sty(ws1.cell(r, ci), bg, "CFD8DC", 10, False, "center", False)

    sep = ws1.max_row + 1
    ws1.row_dimensions[sep].height = 4
    for ci in range(1, ncols1 + 1):
        c = ws1.cell(sep, ci); c.fill = _fill(P["GH"]); c.border = Border(top=_SM, bottom=_SM)

    ld = 4 + len(resumen_final)
    ws1.append([""] + [f"TOTAL - {total_uds} unidades"] +
               [f"=SUM({get_column_letter(ci + 3)}5:{get_column_letter(ci + 3)}{ld})" for ci in range(cantMax)])
    r = ws1.max_row
    ws1.row_dimensions[r].height = 26
    ws1.cell(r, 1).fill = _fill(P["VM"]); ws1.cell(r, 1).border = _BRD
    hdr(ws1.cell(r, 2), P["VM"], "left")
    for ci in range(3, ncols1 + 1):
        hdr(ws1.cell(r, ci), P["VM"])

    ws1.append([f"Generado por FulfillPro - {nowStr}"])
    r = ws1.max_row
    ws1.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols1)
    c = ws1.cell(r, 1)
    c.font = Font(name="Calibri", color="90A4AE", size=8, italic=True)
    c.alignment = _align("left", False)
    ws1.row_dimensions[r].height = 14

    ws1.page_setup.orientation = "portrait"
    ws1.page_setup.paperSize = 9
    ws1.page_setup.fitToPage = True
    ws1.page_setup.fitToWidth = 1
    ws1.page_setup.fitToHeight = 0
    ws1.print_title_rows = "1:4"
    ws1.page_margins = PageMargins(left=0.5, right=0.5, top=0.75, bottom=0.75)

    # ── HOJA 2: REPORTE ORDENADO ──
    ws2 = wb.create_sheet("Reporte Ordenado")
    ws2.sheet_properties.tabColor = P["GH"]
    ws2.sheet_view.showGridLines = False
    ws2.column_dimensions["A"].width = 14
    ws2.column_dimensions["B"].width = 38
    ws2.column_dimensions["C"].width = 10

    banner(ws2, 3, "FulfillPro - Reporte Ordenado", P["VC"], 13, 26)
    subtitle(ws2, 3, f"{len(reporte)} lineas | Fecha: {todayStr}")
    spacer(ws2, 3)

    ws2.sheet_format.defaultRowHeight = 20
    ws2.append(["ID ORDEN", "PRODUCTO", "CANTIDAD"])
    r = ws2.max_row
    ws2.row_dimensions[r].height = 24
    hdr(ws2.cell(r, 1), P["GH"]); hdr(ws2.cell(r, 2), P["GH"], "left"); hdr(ws2.cell(r, 3), P["GH"])
    ws2.freeze_panes = "A5"

    prev = ""
    for i, row in enumerate(reporte):
        isN = row["PRODUCTO"] != prev
        bg  = P["BL"] if i % 2 == 0 else P["GL"]
        ws2.append([str(row["ID ORDEN"]), row["PRODUCTO"], int(row["CANTIDAD"] or 0)])
        r = ws2.max_row
        sty(ws2.cell(r, 1), bg, "546E7A", 9, False, "center", False)
        sty(ws2.cell(r, 2), bg, "212121", 10, isN, "left", True)
        sty(ws2.cell(r, 3), bg, "1565C8", 11, True, "center", False)
        prev = row["PRODUCTO"]

    # ── HOJA 3: PRIORITARIAS ──
    ws3 = wb.create_sheet("PRIORITARIAS")
    ws3.sheet_properties.tabColor = P["RH"]
    ws3.sheet_view.showGridLines = False
    for i, w in enumerate([18, 34, 12, 13, 13, 18, 13], 1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    banner(ws3, 7, "FulfillPro - Ordenes Prioritarias", P["RH"], 13, 28)
    subtitle(ws3, 7, f"{len(prior)} ordenes atrasadas - Riesgo: ${total_riesgo:,.0f} | Fecha: {todayStr}")
    subtitle(ws3, 7, "Rojo intenso = 5+ dias | Rojo suave = 2-4 dias | Naranja = 1 dia")

    ws3.sheet_format.defaultRowHeight = 24
    ws3.append(["N GUIA", "PRODUCTO", "VALOR", "FECHA GUIA", "DIAS RETRASO", "ESTADO", "RIESGO 20%"])
    r = ws3.max_row
    ws3.row_dimensions[r].height = 26
    for ci in range(1, 8):
        hdr(ws3.cell(r, ci), P["RH"])
    hdr(ws3.cell(r, 2), P["RH"], "left")
    ws3.freeze_panes = "A5"

    for row in prior:
        dias = int(row["DIAS RETRASO"])
        if   dias >= 5: bg, fg = "FFCDD2", P["RH"]
        elif dias >= 2: bg, fg = P["RL"], "C62828"
        else:           bg, fg = P["NL"], P["NH"]
        ws3.append([str(row["N GUIA"]), row["PRODUCTO"], float(row["VALOR"]),
                    str(row["FECHA GUIA"]), dias, row["ESTADO"], float(row["RIESGO 20"])])
        r = ws3.max_row
        sty(ws3.cell(r, 1), bg, "546E7A", 9, False, "center", False)
        sty(ws3.cell(r, 2), bg, fg, 10, True, "left", True)
        c3 = ws3.cell(r, 3); sty(c3, bg, "37474F", 10, False, "right", False); c3.number_format = "$#,##0"
        sty(ws3.cell(r, 4), bg, "546E7A", 9, False, "center", False)
        sty(ws3.cell(r, 5), bg, fg, 13, True, "center", False)
        sty(ws3.cell(r, 6), bg, fg, 9, dias >= 2, "center", False)
        c7 = ws3.cell(r, 7); sty(c7, bg, P["RH"], 10, True, "right", False); c7.number_format = "$#,##0"

    if prior:
        sep = ws3.max_row + 1
        for ci in range(1, 8):
            c = ws3.cell(sep, ci); c.fill = _fill(P["GH"]); c.border = Border(top=_SM, bottom=_SM)
        ws3.row_dimensions[sep].height = 4
        lp = 4 + len(prior)
        ws3.append(["", "", "", "", "", "TOTAL RIESGO:", f"=SUM(G5:G{lp})"])
        r = ws3.max_row
        ws3.row_dimensions[r].height = 26
        for ci in range(1, 6):
            c = ws3.cell(r, ci); c.fill = _fill(P["RH"]); c.border = _BRD
        hdr(ws3.cell(r, 6), P["RH"], "right")
        c = ws3.cell(r, 7); hdr(c, P["RH"], "right"); c.number_format = "$#,##0"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
